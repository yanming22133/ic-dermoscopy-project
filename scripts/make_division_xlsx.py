"""生成分工 Excel（中英文并列，多 sheet）。
Generate the division-of-labor Excel (bilingual side-by-side, multiple sheets).

用法 / Usage:
  f:\\anacondaenvs\\pytorch\\python.exe scripts/make_division_xlsx.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(__file__), '..', 'DIVISION_OF_LABOR.xlsx')

HEAD_FILL = PatternFill('solid', fgColor='1F4E78')
HEAD_FONT = Font(bold=True, color='FFFFFF', size=11)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = CENTER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_rows(ws, rows):
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = WRAP


wb = Workbook()

# ---------- Sheet 1: 角色总览 / Roles Overview ----------
ws = wb.active
ws.title = 'Roles 角色总览'
headers = ['No.', 'Role (EN)', '角色 (中文)', 'Claimed by / 认领人',
           'Suitable background (EN)', '适合背景 (中文)', 'Effort / 工作量', 'Code? / 需写代码']
ws.append(headers)
rows = [
    ['R0', 'Tech / experiments lead', '技术/实验总负责', 'Mengzhe Yang (固定/fixed)',
     'Information Mgmt & Info Sys', '信息管理与信息系统', 'Very high 极高', 'Yes (all) 全部'],
    ['R1', 'Literature review lead', '文献综述负责人', '（待认领 to claim）',
     'Applied AI / CS / Math', 'Applied AI / CS / 数学', 'High 高', 'No 否'],
    ['R2', 'Evaluation & metrics lead', '评估与指标负责人', '（待认领 to claim）',
     'Math / Stats / Data Science', '数学 / 统计 / 数据科学', 'Med-high 中高', 'No (Excel ok)'],
    ['R3', 'Data & visualization lead', '数据与可视化负责人', '（待认领 to claim）',
     'Big data / Data analysis', '大数据 / 数据分析', 'Med 中', 'No 否'],
    ['R4', 'Clinical content & report writer', '临床内容+报告主笔', '（待认领 to claim）',
     'Any, strong English writing', '任何背景，英文写作强', 'Med-high 中高', 'No 否'],
    ['R5', 'Project lead & presentation design', '项目负责人+汇报设计', '（待认领 to claim）',
     'Management / Economics / Any', '管理 / 经济 / 任何', 'Med 中', 'No 否'],
]
write_rows(ws, rows)
style_header(ws, len(headers))
set_widths(ws, [5, 28, 22, 22, 28, 22, 16, 16])

# ---------- Sheet 2: 角色详情 / Role Details ----------
ws = wb.create_sheet('Role Details 角色详情')
headers = ['Role (EN)', '角色 (中文)', 'Deliverables (EN)', '产出 (中文)',
           'Q&A owns (EN)', 'Q&A (中文)', 'Deadline']
ws.append(headers)
rows = [
    ['Tech / experiments lead', '技术/实验总负责',
     'All code: Task1/2/3/Bonus training, pipeline, one-click inference; feed raw results to teammates',
     '全部代码：Task1/2/3/Bonus 训练、pipeline、一键推理；产出原始结果喂队友',
     'Model selection, training, ablations, bugs',
     '模型选型、训练细节、消融、bug', '7/29'],
    ['Literature review lead', '文献综述负责人',
     'Read papers/ (30+), write lit review (due 7/27), justify model choice; paired backup: can run inference',
     '读 papers/ 30+篇，写文献综述（7/27交），论证选型；结对备份：能跑推理',
     'Why SegFormer+SAM (from papers read)',
     '为什么选这些模型（读过的论文）', '7/27'],
    ['Evaluation & metrics lead', '评估与指标负责人',
     'Design eval protocol (IoU/per-attr F1/Hausdorff), build result tables, error analysis; design Bonus retrieval-impact comparison',
     '设计评估协议（IoU/逐属性F1/Hausdorff），建结果表，error analysis；设计Bonus retrieval impact对比',
     'Metric meanings, per-attr performance, errors',
     '指标含义、逐属性表现、误差', '7/28'],
    ['Data & visualization lead', '数据与可视化负责人',
     'EDA (class distribution, attribute sparsity), all figures for report+PPT (IoU comparison, confusion matrix, seg samples, retrieval viz)',
     'EDA（类别分布、属性稀疏度），报告+PPT全部图表（IoU对比、混淆矩阵、分割样例、检索可视化）',
     'Data characteristics, figure interpretation',
     '数据特性、图表解读', '7/28'],
    ['Clinical content & report writer', '临床内容+报告主笔',
     '5 terms medical definitions & diagnostic significance; write ≤12-page report (Methods/Results/Discussion/Future); co-design Task3 rules + checklist with Mengzhe',
     '5术语医学释义与诊断意义；主笔≤12页报告；与Mengzhe共同设计Task3规则与一致性checklist',
     'Clinical motivation, terms, discussion & limitations',
     '临床动机、术语、讨论与局限', '7/29'],
    ['Project lead & presentation design', '项目负责人+汇报设计',
     'Timeline, submission & reproducibility checklist, PPT design & narrative, Q&A bank, coordinate daily sync',
     '时间线、提交与复现性checklist、PPT设计与叙事、Q&A题库、协调每日sync',
     'Project flow, overall narrative, conclusion',
     '项目流程、整体叙事、结论', '7/30'],
]
write_rows(ws, rows)
style_header(ws, len(headers))
set_widths(ws, [26, 20, 45, 40, 30, 26, 10])

# ---------- Sheet 3: 汇报分工 / Presentation ----------
ws = wb.create_sheet('Presentation 汇报分工')
headers = ['#', 'Section (EN)', '模块 (中文)', 'Time', 'Speaker / 主讲']
ws.append(headers)
rows = [
    ['1', 'Title, clinical motivation, 5 terms', '题目、临床动机、5术语', '2:00', 'R4'],
    ['2', 'Literature review', '文献综述', '2:30', 'R1'],
    ['3', 'Data analysis (distribution / sparsity)', '数据分析（分布/稀疏度）', '2:00', 'R3'],
    ['4', 'Methods: pipeline + Task1/2/3 + Bonus', '方法：pipeline+Task1/2/3+Bonus', '5:00', 'R0 (Mengzhe)'],
    ['5', 'Results, evaluation, demo', '结果、评估、demo', '2:30', 'R2'],
    ['6', 'Conclusion, future work, team', '结论、未来、致谢团队', '1:00', 'R5'],
    ['', 'Total', '合计', '15:00', ''],
    ['QA', '5 min, all on stage', '5分钟全员站台', '5:00', 'All 全员'],
]
write_rows(ws, rows)
style_header(ws, len(headers))
set_widths(ws, [5, 40, 32, 10, 18])

# ---------- Sheet 4: 接口契约 / Interface Contract ----------
ws = wb.create_sheet('Interface 接口契约')
headers = ['Item (EN)', '项目 (中文)', 'Format / 说明 (bilingual)']
ws.append(headers)
rows = [
    ['Task1 mask', 'Task1 mask',
     '{id}.jpg, L mode, 0/255 (jpg lossy, threshold >127 on read). Metrics: Dice/IoU/95%Hausdorff (tutorial p46). '
     '命名 jpg、L模式、0/255（jpg有损，读取>127还原）。指标 Dice/IoU/95%Hausdorff。'],
    ['Task2 mask', 'Task2 mask',
     '{id}/{attr}.png, L, 0/255; attr = singular milia_like_cyst. '
     '每图一目录，png、0/255；属性名单数 milia_like_cyst。'],
    ['Task2 presence', 'Task2 presence',
     'p_attr = mean(sigmoid(logits)) over lesion ROI (NOT coverage). '
     'p_attr = 病灶ROI上 mean(sigmoid(logits))，不是覆盖率。'],
    ['Task2 status', 'Task2 status',
     'p>=0.60 present / p<=0.40 absent / between uncertain (tutorial p40). '
     'p>=0.60 present / <=0.40 absent / 中间 uncertain。'],
    ['Task3 JSON', 'Task3 JSON',
     'Match example_result schema; attributes_order uses plural milia_like_cysts. '
     '照 example schema；attributes_order 用复数 milia_like_cysts。'],
    ['Task3 thresholds', 'Task3 阈值',
     'border = perimeter²/(4π·area), >=1.60 irregular; size ratio <0.08 small / 0.08-0.25 moderate / >0.25 large. '
     'border=周长²/(4π·面积)>=1.60 irregular；size 比例 <0.08/0.08-0.25/>0.25。'],
    ['Task3 text', 'Task3 文本',
     '"The lesion is {size} with {border} borders. Pigment network is {status}; ..." findings field has a leading space. '
     '"The lesion is {size} with {border} borders..." findings 字段开头有空格。'],
    ['Bonus CSV', 'Bonus CSV',
     'query_image,neighbor_id,similarity. Audit completeness (all imgs, all IDs saved) + sanity (no dup, consistent K). CLIP frozen, no fine-tune; RAG no copy. '
     'audit完整性（全图、ID全存）+ sanity（无重复、K一致）。CLIP冻结不微调；RAG不抄邻居。'],
    ['Preprocessing', '预处理',
     'DullRazor (hair) + Shades of Gray (color constancy) + normalize + resize (tutorial p35). '
     'DullRazor去毛 + Shades of Gray颜色恒常 + 归一化 + resize。'],
    ['Split', '数据划分',
     'Deterministic 80/10/10 train/val/test-local, fixed seed (tutorial p56). Official test set: inference only, never train. '
     '确定性 80/10/10，固定种子。官方测试集只推理不训练。'],
]
write_rows(ws, rows)
style_header(ws, len(headers))
set_widths(ws, [20, 18, 80])

# ---------- Sheet 5: 认领回执 / Signup ----------
ws = wb.create_sheet('Signup 认领回执')
headers = ['No.', 'Role (EN)', '角色 (中文)', 'Claimed by / 认领人', 'Contact / 联系方式']
ws.append(headers)
rows = [
    ['R0', 'Tech / experiments lead', '技术/实验总负责', 'Mengzhe Yang', ''],
    ['R1', 'Literature review lead', '文献综述负责人', '', ''],
    ['R2', 'Evaluation & metrics lead', '评估与指标负责人', '', ''],
    ['R3', 'Data & visualization lead', '数据与可视化负责人', '', ''],
    ['R4', 'Clinical content & report writer', '临床内容+报告主笔', '', ''],
    ['R5', 'Project lead & presentation design', '项目负责人+汇报设计', '', ''],
]
write_rows(ws, rows)
style_header(ws, len(headers))
set_widths(ws, [5, 32, 24, 26, 24])
# 留空认领格高亮
for r in range(2, 8):
    ws.cell(row=r, column=4).fill = PatternFill('solid', fgColor='FFF2CC')

wb.save(OUT)
print('saved', os.path.normpath(OUT))
